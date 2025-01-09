from django.contrib.auth.models import User
from django.core.mail import send_mail
from celery import shared_task
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import pandas as pd
import time
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from datetime import datetime
from django_celery_results.models import TaskResult
from django_celery_beat.models import PeriodicTask
import openpyxl
from openpyxl.styles import Alignment
from openpyxl import Workbook
from core.models import *
import os
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.core.mail import EmailMessage

@shared_task
def scrape_to_excel():

    start_time = datetime.now()
    datos_obtenidos = []
    admin_email = User.objects.filter(is_superuser=True).first().email  # Obtén el correo del admin

    # Conjuntos de datos para las búsquedas
    data_sets = [
        {"rut": "99531960", "dv": "8", "era": "2024", "competencia": "Laboral"},
        {"rut": "79883910", "dv": "1", "era": "2024", "competencia": "Laboral"},
        {"rut": "76409376", "dv": "3", "era": "2024", "competencia": "Laboral"},
    ]

    try:
        # Configura el driver
        driver = webdriver.Chrome()  # Asegúrate de tener `chromedriver` configurado en tu sistema
        
        modal_cerrado = False
        consulta_causas_click = False

        datos_por_rut = {data["rut"]: [] for data in data_sets}
        for data in data_sets:
            try:
                # Navegar a la página principal
                driver.get("https://oficinajudicialvirtual.pjud.cl/indexN.php")
                time.sleep(5)

                # Manejar el modal solo la primera vez
                if not modal_cerrado:
                    try:
                        modal = WebDriverWait(driver, 10).until(
                            EC.visibility_of_element_located((By.ID, "no-disponible"))
                        )
                        boton_cerrar = WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable((By.CSS_SELECTOR, "#no-disponible button[data-dismiss='modal']"))
                        )
                        boton_cerrar.click()
                        modal_cerrado = True  # Actualizar bandera
                    except TimeoutException:
                        modal_cerrado = True  # Considerar que no aparece el modal

                # Hacer clic en "Consulta causas" solo la primera vez
                if not consulta_causas_click:
                    driver.find_element(By.XPATH, "//button[text()='Consulta causas']").click()
                    consulta_causas_click = True  # Actualizar bandera

                time.sleep(3)

                # Navegar a la sección de búsqueda jurídica
                driver.find_element(By.XPATH, "//a[@href='#BusJuridica']").click()
                time.sleep(3)

                # Rellenar formulario
                driver.find_element(By.ID, "rutJur").send_keys(data["rut"])
                driver.find_element(By.ID, "dvJur").send_keys(data["dv"])
                driver.find_element(By.ID, "eraJur").send_keys(data["era"])
                driver.find_element(By.ID, "jurCompetencia").send_keys(data["competencia"])
                time.sleep(3)

                borrar = [["No se han encontrado resultados con los datos ingresados. Recuerde que las causas reservadas no se muestran en la consulta unificada, y según el tipo de reserva, podrá acceder a ellas ingresando con su usuario y contraseña a la opción “Mis Causas” de la Oficina Judicial Virtual. Para conocer los tipos de reserva, pinchar aquí"]]
                

                rut_actual = data["rut"]
                if rut_actual not in datos_por_rut:
                    datos_por_rut[rut_actual] = []
                # Seleccionar cada tribunal
                for tribunal_index in range(147):
                    try:
                        # Selección de tribunal
                        combobox = driver.find_element(By.ID, "jurTribunal")
                        select = Select(combobox)
                        select.select_by_index(tribunal_index + 2)
                        driver.find_element(By.ID, "btnConConsultaJur").click()
                        time.sleep(6)

                        # Extraer tabla
                        tabla = WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.ID, "dtaTableDetalleJuridica"))
                        )
                        filas = tabla.find_elements(By.TAG_NAME, "tr")

                        for fila in filas:
                            columnas = fila.find_elements(By.TAG_NAME, "td")
                            if len(columnas) > 5:  # Verificar que hay suficientes columnas
                                # Ignorar la primera columna y procesar solo las 5 siguientes
                                datos_fila = [columna.text.strip() for columna in columnas[1:6]]
                                if datos_fila not in borrar:  # Verificar que no sea el mensaje de exclusión
                                    datos_por_rut[rut_actual].append(datos_fila)
                    except Exception as e:
                        print(f"Error procesando tribunal {tribunal_index} para el RUT {rut_actual}: {e}")

            except Exception as e:
                print(f"Error con el conjunto de datos {data}: {e}")
                continue  # Continuar con el siguiente conjunto de datos

    except Exception as e:
        # Enviar correo en caso de error crítico
        error_message = f"La tarea falló. Error: {e}\nFecha y hora: {datetime.now()}"
        send_mail(
            subject="Error en tarea asincrónica: scrape_to_excel",
            message=error_message,
            from_email="diegitocool@gmail.com",
            recipient_list=[admin_email],
            fail_silently=False,
        )
        driver.quit()
        raise e

    finally:
        driver.quit()

    # Crear un libro de trabajo y una hoja
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Resultados"

    # Inicializar fila de escritura (comienza después de la fila en blanco inicial)
    current_row = 2  # Dejar la primera fila en blanco

    # Escribir datos por RUT
    for rut, datos in datos_por_rut.items():
        if not datos:
            print(f"Sin datos para el RUT: {rut}")
            continue

        # Escribir encabezados con RUT como primera columna
        headers = ["RUT", "Rit", "Tribunal", "Caratulado", "Fecha Ingreso", "Estado Causa"]
        for col_index, header in enumerate(headers, start=1):  # Comienza en la primera columna
            sheet.cell(row=current_row, column=col_index, value=header)
        current_row += 1  # Mover a la fila siguiente

        # Escribir datos con RUT en la primera columna
        for row_data in datos:
            sheet.cell(row=current_row, column=1, value=rut)  # Escribir el RUT en la primera columna
            for col_index, cell_data in enumerate(row_data, start=2):  # Continuar desde la segunda columna
                sheet.cell(row=current_row, column=col_index, value=cell_data)
            current_row += 1  # Mover a la fila siguiente

        # Agregar una fila en blanco entre tablas
        current_row += 1

    # Crear la carpeta si no existe
    folder_name = "excels_judicial"
    os.makedirs(os.path.join("media", folder_name), exist_ok=True)

    # Generar un nombre dinámico para el archivo
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    file_name = f"Resultados_Buscados_{timestamp}.xlsx"
    file_path = os.path.join("media", folder_name, file_name)

    # Guardar el archivo Excel
    workbook.save(file_path)
    print(f"Archivo Excel generado correctamente: {file_path}")

    # Crear un registro en InfoScrap
    info_scrap = InfoScrap(documento=os.path.join(folder_name, file_name))
    info_scrap.save()
    print(f"Archivo guardado en el modelo InfoScrap: {info_scrap.documento.url}")

    compare_excel()
    return {
        "message": "Scraping completado y función activada.",
    }

def compare_excel():
    print("Se activó la función secundaria.")
    admin_email = User.objects.filter(is_superuser=True).first().email

    # Obtener los dos últimos registros en InfoScrap
    last_two_excels = InfoScrap.objects.order_by('-fecha_guardado')[:2]

    if len(last_two_excels) < 2:
        print("No hay suficientes archivos para comparar.")
        return

    # Extraer los archivos
    latest_file = last_two_excels[0].documento.path
    previous_file = last_two_excels[1].documento.path

    try:
        # Cargar los archivos Excel con pandas sin encabezados y asignar nombres manualmente
        latest_df = pd.read_excel(latest_file, engine='openpyxl', header=None)
        previous_df = pd.read_excel(previous_file, engine='openpyxl', header=None)

        # Asignar nombres a las columnas
        column_names = ['Rut', 'Rit', 'Tribunal', 'Caratulado', 'Fecha Ingreso', 'Estado Causa']
        latest_df.columns = column_names
        previous_df.columns = column_names

        # Asegurar que ambos DataFrames tengan el mismo número de filas
        max_rows = min(len(latest_df), len(previous_df))

        # Detectar diferencias en columnas (valores cambiados)
        column_diffs = latest_df.iloc[:max_rows].ne(previous_df.iloc[:max_rows])  # Compara celda por celda
        changed_cells = column_diffs.stack()  # Obtiene las celdas que cambiaron
        changed_cells = changed_cells[changed_cells]  # Filtra solo las celdas con cambios

        # Filtrar celdas cuyo valor anterior y nuevo son `nan` o vacíos
        filtered_changes = []
        for (index, column), _ in changed_cells.items():
            old_value = previous_df.at[index, column]
            new_value = latest_df.at[index, column]
            if not (pd.isna(old_value) and pd.isna(new_value)):  # Excluye cambios entre valores vacíos
                filtered_changes.append({
                    'row': index + 2,  # Ajusta para que la fila comience en 1 (estilo Excel)
                    'column': column,
                    'old_value': old_value,
                    'new_value': new_value,
                })

        # Detectar filas nuevas en el archivo más reciente
        new_rows = latest_df.iloc[max_rows:].to_dict(orient='records')

        # Si no hubo cambios relevantes ni nuevas filas
        if not filtered_changes and not new_rows:
            print("No se encontraron diferencias relevantes ni nuevas filas entre los dos últimos archivos Excel.")
            
            # Renderizar plantilla para el caso sin diferencias
            email_html_content = render_to_string(
                'no_differences_email.html',
                {
                    'message': "No se encontraron diferencias relevantes ni nuevas filas entre los dos últimos archivos Excel."
                }
            )
            
            # Crear y enviar el correo
            email = EmailMessage(
                subject="Sin diferencias en la comparación de archivos Excel",
                body=email_html_content,
                from_email="diegitocool@gmail.com",
                to=[admin_email],
            )
            email.content_subtype = "html"  # Definir el contenido como HTML
            email.send()
            return

        # Renderizar la plantilla HTML para el caso con diferencias
        email_html_content = render_to_string(
            'comparison_email.html',
            {
                'changed_data': filtered_changes,
                'new_rows': new_rows,
            }
        )

        # Crear y enviar el correo
        email = EmailMessage(
            subject="Resultado de comparación entre archivos Excel",
            body=email_html_content,
            from_email="diegitocool@gmail.com",
            to=[admin_email],
        )
        email.content_subtype = "html"  # Definir el contenido como HTML
        email.send()

        print("Correo enviado con éxito.")

    except Exception as e:
        print(f"Error al comparar los archivos Excel: {e}")